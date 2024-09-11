import pandas as pd
import numpy as np
import shutil
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def read_folder_and_create_excel(directory):
    # Put all the folder names in the directory into a list
    folders = [folder for folder in os.listdir(directory) if os.path.isdir(os.path.join(directory, folder)) and '_duplicate' in folder]
    
    # Extract the temperature and current setting and save them in a dataframe
    for folder in folders:
        temperatures = []
        currents = []

        if 'C' in folder and 'A' in folder:
            temp_index = folder.index('C')
            curr_index = folder.index('A')

            try:
                temperature = float(folder[:temp_index])
                current = float(folder[temp_index+1:curr_index])
                temperatures.append(temperature)
                currents.append(current)
            except ValueError:
                continue
            
        filenumber = len([name for name in os.listdir(os.path.join(directory, folder)) if os.path.isfile(os.path.join(directory, folder, name))])

        df = pd.DataFrame({
            'Set Temperature (C)': temperatures * filenumber,
            'Set Current (A)': currents * filenumber
        })

        excel_path = os.path.join(directory, folder, f"{folder}.xlsx")

        # Write to excel file created
        if not os.path.exists(excel_path):
            df.to_excel(excel_path, index=False)
        else:
            print("Excel file already exists.")
    
    

def duplicate_and_rename_folder(directory):
    folders = [folder for folder in os.listdir(directory) if os.path.isdir(os.path.join(directory, folder)) and 'A' in folder and 'C' in folder and '_duplicate' not in folder]
    for folder in folders:
        source_folder_path = os.path.join(directory, folder)
        target_folder_path = os.path.join(directory, folder +'_duplicates')
        os.makedirs(target_folder_path, exist_ok=True)

        for filename in os.listdir(source_folder_path):
            original_file = os.path.join(source_folder_path, filename)
            duplicated_file = os.path.join(target_folder_path, filename)
            shutil.copy(original_file, duplicated_file)

            if filename.endswith('.all'):
                txt_filename = filename[:-4] + '.txt'
                txt_file_path = os.path.join(target_folder_path, txt_filename)
                shutil.move(duplicated_file, txt_file_path)
        
def read_and_update_excel(directory):
    folders = [folder for folder in os.listdir(directory) if os.path.isdir(os.path.join(directory, folder)) and '_duplicates' in folder]
    for folder in folders:
        folder_path = os.path.join(directory, folder)
        excel_path = os.path.join(folder_path, f"{folder}.xlsx")

        if os.path.exists(excel_path):
            df = pd.read_excel(excel_path)
        else:
            read_folder_and_create_excel(directory)
            df = pd.read_excel(excel_path)
        columns = ['time (s)']
        measurements = ['actual temperature (C)']
        for ch in range(1, 7):
            measurements +=[f'CH{ch} actual current (A)', f'CH{ch} actual Voltage', f'CH{ch} resistance (ohm)']
        columns.extend(measurements)

        new_data = pd.DataFrame(columns=columns)

        for filename in os.listdir(folder):
            if filename.endswith('.txt'):
                file_path = os.path.join(directory, folder, filename)
                with open(file_path, 'r') as file:
                    lines = [line.strip().split() for line in file.readlines()[:7]]
                    data = np.array(lines, dtype=float)

                    time = data[0, 0]
                    temperature = data[0, 1]
                    currents = data[1:, 6]
                    volatges = data[1:, 0]
                    resistances = volatges / currents

                    row = [time, temperature] + [val for pair in zip(currents, volatges, resistances) for val in pair]
                    new_data.loc[len(new_data)] = row
        
        if 'df' in locals():
            # full_df = pd.concat([df, new_data], axis=1, ignore_index=False)
            full_df = pd.merge(df, new_data, left_index=True, right_index=True)
            full_df.to_excel(excel_path, index=False)
            path = Path(excel_path)
            new_file_name = path.stem.replace('_duplicates', '') + path.suffix
            # Define the destination path (moving the file to its parent directory)
            destination_path = path.parent.parent / new_file_name

            # Move the file to its parent directory
            shutil.move(str(path), str(destination_path))
        else:
            new_data.to_excel(excel_path, index=False)
        
        shutil.rmtree(os.path.join(directory, folder))

def combine_excel_files(directory, output_filename):
    excel_files = [os.path.join(dp, f) for dp, dn, filename in os.walk(directory) for f in filename if f.endswith('.xlsx')]
    df_list = [pd.read_excel(file) for file in excel_files]
    combined_df = pd.concat(df_list, ignore_index=True)
    combined_df.to_excel(output_filename, index=False)

def remove_other_excel_files(directory, keep_files):
    for dp, dn, filenames in os.walk(directory):
        for file in filenames:
            if file.endswith('.xlsx') and file != keep_files:
                os.remove(os.path.join(dp, file))

def add_sheet_excel(directory, excel_name):

    # Put all the folder names in the directory into a list
    path = os.path.join(directory, excel_name)
    folders = [folder for folder in os.listdir(directory) if os.path.isdir(os.path.join(directory, folder))]
    temp_values = []
    current_values = []
    # Extract the temperature and current setting and save them in a dataframe
    for folder in folders:

        if 'C' in folder and 'A' in folder:
            temp_index = folder.index('C')
            curr_index = folder.index('A')

            try:
                temperature = float(folder[:temp_index])
                current = float(folder[temp_index+1:curr_index])
                temp_values.append(temperature)
                current_values.append(current)
            except ValueError:
                continue

    T = sorted(list(set(temp_values)))
    I = sorted(list(set(current_values)))

    index = pd.MultiIndex.from_product([range(1, 7), T], names=["Channel", "T (C)\\I (A)"])
    df = pd.DataFrame(index=index, columns=I)
    
    with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name='Sheet2')
    
    wb = load_workbook(path)
    ws = wb['Sheet2']

    start_row = 2
    for ch in range(1, 7):
        end_row = start_row + len(T) -1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.cell(row=start_row, column=1, value=f'CH{ch}')
        start_row = end_row + 1

    wb.save(path)
    wb.close()

def sort_column (directoy, excel_name):
    target_column = 0  
    path = os.path.join(directoy, excel_name)
    # Load the existing workbook
    workbook = load_workbook(path)
    sheet = workbook.active  # Get the active sheet

    # Read all data from the sheet
    data = list(sheet.iter_rows(values_only=True))

    # Separate headers and data
    labels = data[0]    # Don't sort the headers
    data = data[1:]     # Data begins on the second row

    # Sort data by the target column
    data.sort(key=lambda x: x[target_column])

    # Write sorted data back into the same sheet
    for idx, label in enumerate(labels):
        sheet.cell(row=1, column=idx+1, value=label)

    for idx_r, row in enumerate(data):
        for idx_c, value in enumerate(row):
            sheet.cell(row=idx_r+2, column=idx_c+1, value=value)

    # Save the modified workbook back to the same file
    workbook.save(path)

def main():
    directory_path = os.getcwd()
    duplicate_and_rename_folder(directory_path)
    read_folder_and_create_excel(directory_path)
    read_and_update_excel(directory_path)
    combine_excel_files(directory_path, 'Result.xlsx')
    sort_column (directory_path, 'Result.xlsx')
    remove_other_excel_files(directory_path, 'Result.xlsx')
    add_sheet_excel(directory_path, 'Result.xlsx')


if __name__ == "__main__":
    main()
