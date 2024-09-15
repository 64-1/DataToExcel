from openpyxl import load_workbook
import os

def update_resistance_values(directory, excel_name):
    path = os.path.join(directory, excel_name)
    wb = load_workbook(path)
    ws = wb.active  # Assuming the data is in the active sheet

    # Initialize column indices
    col_indices = {
        "time": None,
        "Set Temperature": None,
        "Set Current": None,
        "CH1 resistance": None,
        "CH2 resistance": None,
        "CH3 resistance": None,
        "CH4 resistance": None,
        "CH5 resistance": None,
        "CH6 resistance": None
    }

    # Find all relevant columns in the header row
    for cell in ws[1]:  # ws[1] accesses the first row
        for key in col_indices:
            if cell.value and key.lower() in str(cell.value).lower():
                col_indices[key] = cell.column

    # Debug to check column indices
    print("Column Indices:", col_indices)

    # Build a dictionary to store the maximum time for each (Set Temperature, Set Current)
    max_time_rows = {}  # Key: (Set Temperature, Set Current), Value: (max_time, row_number)

    # Iterate over all data rows to find the maximum time for each combination
    for row_number in range(2, ws.max_row + 1):
        try:
            set_temp = float(ws.cell(row=row_number, column=col_indices["Set Temperature"]).value)
            set_current = float(ws.cell(row=row_number, column=col_indices["Set Current"]).value)
            time_value = float(ws.cell(row=row_number, column=col_indices["time"]).value)
        except (TypeError, ValueError):
            continue  # Skip rows with invalid data

        key = (set_temp, set_current)
        if key not in max_time_rows or time_value > max_time_rows[key][0]:
            max_time_rows[key] = (time_value, row_number)

    # Open Sheet2
    ws_sheet2 = wb['Sheet2']

    # Get the currents from the header row of Sheet2
    currents = []
    for col in range(3, ws_sheet2.max_column + 1):
        current_value = ws_sheet2.cell(row=1, column=col).value
        try:
            currents.append(float(current_value))
        except (TypeError, ValueError):
            continue  # Skip if the value cannot be converted to float

    # Process each row with the maximum time
    for (set_temp, set_current), (max_time, row_number) in max_time_rows.items():
        data = {
            "Set Temperature": set_temp,
            "Set Current": set_current,
        }
        for i in range(1, 7):  # For each channel
            resistance_key = f"CH{i} resistance"
            resistance_value = ws.cell(row=row_number, column=col_indices[resistance_key]).value
            data["Channel"] = f"CH{i}"
            data["Resistance"] = resistance_value
            fill_sheet2(ws_sheet2, currents, data)

    wb.save(path)
    wb.close()
