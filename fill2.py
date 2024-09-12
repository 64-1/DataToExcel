from openpyxl import load_workbook

def fill_sheet2(directory, excel_name, data):
    path = os.path.join(directory, excel_name)
    wb = load_workbook(path)
    sheet_name = 'Sheet2'

    # Check if 'Sheet2' exists, if not create it
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # If creating a new sheet, set headers at specific location (Row 1, Column 3 to 6)
        headers = list(data.keys())
        for idx, header in enumerate(headers, start=3):  # Start from column C
            ws.cell(row=1, column=idx, value=header)
    else:
        ws = wb[sheet_name]

    # Find the next available row in the range from row 2 to 25
    # If row 25 is already filled, you will need to decide how to handle it (e.g., overwrite, skip, or throw error)
    next_row = 2  # Start from row 2
    while next_row <= 25 and ws.cell(row=next_row, column=3).value is not None:
        next_row += 1
    
    if next_row > 25:
        print("Error: Rows are already filled up to row 25.")
        return  # Or handle this case as needed

    # Set data starting from column 3
    data_values = list(data.values())
    for idx, value in enumerate(data_values, start=3):  # Start from column C
        ws.cell(row=next_row, column=idx, value=value)

    # Save the workbook
    wb.save(path)
    wb.close()
