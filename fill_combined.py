from openpyxl import load_workbook
import os


def update_resistance_values(directory, excel_name, target_time=3600, tolerance=10):
    path = os.path.join(directory, excel_name)
    wb = load_workbook(path)
    ws = wb.active  # Assuming the data is in the active sheet

    # Initialize column indices
    col_indices = {
        "time": None,
        "Set Temperature": None,
        "Set Current": None,
        "CH1 resistance": None,
        "CH2 resistance": None,
        "CH3 resistance": None,
        "CH4 resistance": None,
        "CH5 resistance": None,
        "CH6 resistance": None
    }

    # Find all relevant columns in the header row
    for cell in ws[1]:  # ws[1] accesses the first row
        for key in col_indices:
            if cell.value and key.lower() in str(cell.value).lower():
                col_indices[key] = cell.column

    # Debug to check column indices
    print("Column Indices:", col_indices)

    # Open Sheet2
    ws_sheet2 = wb['Sheet2']

    # Get the currents from the header row of Sheet2
    currents = []
    for col in range(3, ws_sheet2.max_column + 1):
        current_value = ws_sheet2.cell(row=1, column=col).value
        try:
            currents.append(float(current_value))
        except (TypeError, ValueError):
            continue  # Skip if the value cannot be converted to float

    # Process all rows where time is approximately target_time
    for row_number in range(2, ws.max_row + 1):
        time_cell_value = ws.cell(row=row_number, column=col_indices["time"]).value
        try:
            time_value = float(time_cell_value)
        except (TypeError, ValueError):
            continue  # Skip rows with invalid time values

        if target_time - tolerance <= time_value <= target_time + tolerance:
            for i in range(1, 7):  # For each channel
                resistance_key = f"CH{i} resistance"
                data = {
                    "Channel": f"CH{i}",
                    "Set Temperature": ws.cell(row=row_number, column=col_indices["Set Temperature"]).value,
                    "Set Current": ws.cell(row=row_number, column=col_indices["Set Current"]).value,
                    "Resistance": ws.cell(row=row_number, column=col_indices[resistance_key]).value
                }
                print(f"Data for {resistance_key}: ", data)
                fill_sheet2(ws_sheet2, currents, data)

    wb.save(path)
    wb.close()

def fill_sheet2(ws_sheet2, currents, data):
    # Convert the Set Current to float for comparison
    try:
        set_current = float(data["Set Current"])
    except (TypeError, ValueError):
        print(f"Invalid Set Current: {data['Set Current']}")
        return

    # Find the column index for the Set Current
    try:
        current_index = currents.index(set_current) + 3  # +3 because currents start from column 3
    except ValueError:
        print(f"Set Current {set_current} not found in currents.")
        return

    # Find the row that matches the Channel and Set Temperature
    found_row = None
    for row in range(2, ws_sheet2.max_row + 1):
        channel_value = ws_sheet2.cell(row=row, column=1).value
        temperature_value = ws_sheet2.cell(row=row, column=2).value

        # Ensure temperature values are comparable as floats
        try:
            temp_value = float(temperature_value)
            set_temp = float(data["Set Temperature"])
        except (TypeError, ValueError):
            continue  # Skip rows with invalid temperature values

        if channel_value == data["Channel"] and temp_value == set_temp:
            found_row = row
            break

    if found_row is None:
        print(f"Row with Channel {data['Channel']} and Temperature {data['Set Temperature']} not found.")
        return

    # Write the Resistance value into the correct cell
    ws_sheet2.cell(row=found_row, column=current_index, value=data["Resistance"])



# Example usage:
directory = os.getcwd()
excel_name = 'Result.xlsx'
update_resistance_values(directory, excel_name)
