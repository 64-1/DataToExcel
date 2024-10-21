import os
import sys
import numpy as np
import pandas as pd
import re
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QVBoxLayout

def read_data_from_folders(directory):
    import re

    # Regular expression to match folder names like '25C5A'
    folder_pattern = re.compile(r'([\d.]+)C([\d.]+)A')

    folders = [folder for folder in os.listdir(directory) if os.path.isdir(os.path.join(directory, folder))]
    all_data = []
    for folder in folders:
        folder_path = os.path.join(directory, folder)

        # Use regex to extract temperature and current
        match = folder_pattern.search(folder)
        if match:
            temperature = float(match.group(1))
            current = float(match.group(2))
        else:
            print(f"Folder name {folder} does not match pattern.")
            continue

        print(f"Processing folder: {folder}, Temperature: {temperature}, Current: {current}")

        total_rows_added = 0  # Initialize counter for this folder

        # Process all .all files in the folder
        for filename in os.listdir(folder_path):
            if filename.endswith('.all'):
                print(f"Found .all file: {filename} in folder {folder}")
                file_path = os.path.join(folder_path, filename)
                with open(file_path, 'r') as file:
                    lines = [line.strip().split() for line in file.readlines()]
                    print(f"Read {len(lines)} lines from {filename}")
                    if not lines:
                        print(f"No lines read from {filename}")
                        continue
                    else:
                        print(f"First few lines from {filename}: {lines[:3]}")

                    try:
                        data = np.array(lines, dtype=float)
                    except ValueError as e:
                        print(f"Could not convert data in file {file_path} to float. Error: {e}")
                        continue

                    print(f"Data shape: {data.shape}")
                    print(f"Data: {data}")

                    # Check if data dimensions are as expected
                    if data.shape[0] < 2 or data.shape[1] < 7:
                        print(f"Data in {filename} does not have expected dimensions.")
                        continue

                    time = data[0, 0]
                    set_temperature = temperature
                    set_current = current
                    actual_temperature = data[0, 1]

                    # Extract currents and voltages
                    # Assuming currents are in column 6, voltages are in column 0
                    currents = data[1:, 6]
                    voltages = data[1:, 0]
                    resistances = voltages / currents

                    print(f"Currents: {currents}")
                    print(f"Voltages: {voltages}")
                    print(f"Resistances: {resistances}")

                    # Identify valid channels (where currents or voltages are non-zero)
                    valid_indices = []
                    for i in range(len(currents)):
                        if currents[i] != 0 and voltages[i] != 0:
                            valid_indices.append(i)
                        else:
                            # Stop processing further channels once a zero value is encountered
                            break

                    num_channels = len(valid_indices)
                    if num_channels == 0:
                        print(f"No valid channels found in {filename}")
                        continue

                    # Create a row with all valid channel data
                    row = {
                        'time': time,
                        'Set Temperature': set_temperature,
                        'Set Current': set_current,
                        'actual temperature': actual_temperature,
                    }

                    for idx, i in enumerate(valid_indices):
                        ch = idx + 1  # Channel numbering starts from 1
                        row[f'Actual Current CH{ch}'] = currents[i]
                        row[f'Actual Voltage CH{ch}'] = voltages[i]
                        row[f'Resistance CH{ch}'] = resistances[i]

                    all_data.append(row)
                    total_rows_added += 1
            else:
                print(f"Skipping non-.all file: {filename} in folder {folder}")

        print(f"Total rows added from folder {folder}: {total_rows_added}")

    print(f"Total rows collected: {len(all_data)}")
    df = pd.DataFrame(all_data)
    return df

def sort_column(directory, excel_name):
    target_column = 1  # Assuming 'Set Temperature' is the second column (index 1)
    path = os.path.join(directory, excel_name)
    # Load the existing workbook
    workbook = load_workbook(path)
    sheet = workbook.active  # Get the active sheet

    # Read all data from the sheet
    data = list(sheet.iter_rows(values_only=True))

    if not data:
        print("No data found in the sheet.")
        return

    # Separate headers and data
    labels = data[0]    # Don't sort the headers
    data = data[1:]     # Data begins on the second row

    # Sort data by the target column (Set Temperature)
    data.sort(key=lambda x: x[target_column])

    # Write sorted data back into the same sheet
    for idx_c, label in enumerate(labels):
        sheet.cell(row=1, column=idx_c+1, value=label)

    for idx_r, row in enumerate(data):
        for idx_c, value in enumerate(row):
            sheet.cell(row=idx_r+2, column=idx_c+1, value=value)

    # Save the modified workbook back to the same file
    workbook.save(path)

def add_sheet_excel(directory, excel_name):
    path = os.path.join(directory, excel_name)
    # Read the main data
    df = pd.read_excel(path)

    # Get unique temperatures and currents
    T = sorted(df['Set Temperature'].unique())
    I = sorted(df['Set Current'].unique())

    # Determine the number of channels from the DataFrame columns
    resistance_cols = [col for col in df.columns if col.startswith('Resistance CH')]
    channels = [col.replace('Resistance ', '') for col in resistance_cols]
    num_channels = len(channels)

    # Create a workbook and add Sheet2
    wb = load_workbook(path)
    if 'Sheet2' in wb.sheetnames:
        ws = wb['Sheet2']
    else:
        ws = wb.create_sheet('Sheet2')

    start_row = 1
    for ch in channels:
        ch_str = ch
        # Add Channel label
        ws.cell(row=start_row, column=1, value=ch_str)

        # Prepare data for this channel
        ch_data = df[['Set Temperature', 'Set Current', f'Resistance {ch}']]

        # Pivot the data so that temperatures are rows and currents are columns
        pivot_table = ch_data.pivot_table(values=f'Resistance {ch}', index='Set Temperature', columns='Set Current')

        # Sort the index and columns
        pivot_table = pivot_table.reindex(index=T, columns=I)

        # Convert pivot_table to list of lists
        # First row: header with currents
        header = ['T (C)\\I (A)'] + [str(i) for i in I]
        data_rows = [header]
        for temp in T:
            row = [temp]
            for current in I:
                value = pivot_table.loc[temp, current] if (temp in pivot_table.index and current in pivot_table.columns) else None
                row.append(value)
            data_rows.append(row)

        # Write the data
        for i, row_data in enumerate(data_rows):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i + 1, column=j + 1, value=value)

        # Move to the next channel, adding 2 empty rows
        start_row = start_row + len(data_rows) + 2

    # Save the workbook
    wb.save(path)
    wb.close()

def update_resistance_values(directory, excel_name):
    path = os.path.join(directory, excel_name)
    wb = load_workbook(path)
    ws_sheet2 = wb['Sheet2']

    # Read the main data
    df = pd.read_excel(path)

    # For each Set Temperature and Set Current, find the row with maximum time
    grouped = df.groupby(['Set Temperature', 'Set Current'])

    max_time_rows = grouped.apply(lambda x: x.loc[x['time'].idxmax()])

    # Build a mapping of channel to starting row in Sheet2
    channel_data_info = {}
    row_num = 1
    while row_num <= ws_sheet2.max_row:
        cell_value = ws_sheet2.cell(row=row_num, column=1).value
        if cell_value and isinstance(cell_value, str) and cell_value.startswith('CH'):
            channel = cell_value
            # Data starts two rows below the channel label
            data_start_row = row_num + 2
            # Find the end of the data block
            data_end_row = data_start_row
            while data_end_row <= ws_sheet2.max_row:
                temp_cell_value = ws_sheet2.cell(row=data_end_row, column=1).value
                if temp_cell_value is None or (isinstance(temp_cell_value, str) and temp_cell_value.startswith('CH')):
                    break
                data_end_row += 1
            # Record the data range for this channel
            channel_data_info[channel] = {
                'currents_row': data_start_row - 1,
                'data_start_row': data_start_row,
                'data_end_row': data_end_row - 1
            }
            # Continue from the end of this data block
            row_num = data_end_row
        else:
            row_num += 1

    T = sorted(df['Set Temperature'].unique())
    I = sorted(df['Set Current'].unique())

    # Get the list of channels from the DataFrame
    resistance_cols = [col for col in df.columns if col.startswith('Resistance CH')]
    channels = [col.replace('Resistance ', '') for col in resistance_cols]

    for ch in channels:
        ch_str = ch
        if ch_str not in channel_data_info:
            print(f"Channel {ch_str} not found in Sheet2.")
            continue
        data_info = channel_data_info[ch_str]
        currents_row = data_info['currents_row']
        data_start_row = data_info['data_start_row']
        data_end_row = data_info['data_end_row']

        # Get the mapping of currents to columns
        currents = {}
        for col in range(2, ws_sheet2.max_column + 1):
            current_value = ws_sheet2.cell(row=currents_row, column=col).value
            try:
                current_value = float(current_value)
                currents[current_value] = col
            except (TypeError, ValueError):
                continue

        # Build a mapping of temperatures to rows
        temperatures = {}
        for row in range(data_start_row, data_end_row + 1):
            temperature_value = ws_sheet2.cell(row=row, column=1).value
            try:
                temp_value = float(temperature_value)
                temperatures[temp_value] = row
            except (TypeError, ValueError):
                continue

        # Now, for each (Set Temperature, Set Current), update the resistance value
        for (set_temp, set_current), row_data in max_time_rows.iterrows():
            resistance_value = row_data.get(f'Resistance {ch}', None)

            if resistance_value is None:
                continue

            if set_current not in currents:
                print(f"Set Current {set_current} not found in Sheet2 for {ch_str}.")
                continue
            current_col = currents[set_current]

            if set_temp not in temperatures:
                print(f"Temperature {set_temp} not found in Sheet2 for {ch_str}.")
                continue
            found_row = temperatures[set_temp]

            # Write the Resistance value into the correct cell
            ws_sheet2.cell(row=found_row, column=current_col, value=resistance_value)

    wb.save(path)
    wb.close()

# PyQt5 GUI for folder selection
class FolderSelector(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.selected_directory = None

    def initUI(self):
        self.setWindowTitle('Select Folder for Operation')
        self.setGeometry(100, 100, 400, 100)

        self.layout = QVBoxLayout()

        self.btn_select_folder = QPushButton('Select Folder', self)
        self.btn_select_folder.clicked.connect(self.open_folder_dialog)
        self.layout.addWidget(self.btn_select_folder)

        self.setLayout(self.layout)

    def open_folder_dialog(self):
        folder = QFileDialog.getExistingDirectory(self, 'Select Folder')
        if folder:
            self.selected_directory = folder
            self.close()  # Close the GUI after selection

def main(directory_path):
    print("Reading data from folders...")
    df = read_data_from_folders(directory_path)
    print("Data read successfully.")
    print(df.head())
    print(f"DataFrame shape: {df.shape}")

    if df.empty:
        print("No data was read from the folders.")
        return

    # Write data to 'Result.xlsx'
    output_path = os.path.join(directory_path, 'Result.xlsx')
    df.to_excel(output_path, index=False)
    print(f"Data written to {output_path}")

    # Sort the data
    sort_column(directory_path, 'Result.xlsx')
    print("Data sorted by Set Temperature.")

    # Add Sheet2 and format it
    add_sheet_excel(directory_path, 'Result.xlsx')
    print("Sheet2 added and formatted.")

    # Update resistance values in Sheet2
    update_resistance_values(directory_path, 'Result.xlsx')
    print("Resistance values updated in Sheet2.")

    print("Processing completed.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    selector = FolderSelector()
    selector.show()
    app.exec_()

    # After the GUI is closed, check if a directory was selected
    if selector.selected_directory:
        selected_directory = selector.selected_directory
        main(selected_directory)
    else:
        print("No folder was selected.")
